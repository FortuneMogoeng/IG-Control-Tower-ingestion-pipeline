"""
parse_excel.py — Excel (.xlsx / .xlsm) parser for AI Control Tower ingestion.
Reads each sheet and groups consecutive non-empty rows into fragments,
preserving the sheet name and row range in the provenance location.

Usage:
    python parse_excel.py <path/to/file.xlsx>

Output (stdout, one JSON object per line):
    {
        "fragment_id": "uuid4",
        "source_file": "filename.xlsx",
        "file_type": "xlsx",
        "location": "Sheet 'Summary' rows 3-18",
        "text": "Header | Value | Notes\nRevenue | £1.2m | FY2024\n...",
        "char_count": 289
    }

Requires: openpyxl  (pip install openpyxl)
"""

import json
import sys
import uuid
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("openpyxl not installed. Run: pip install openpyxl")


MIN_CHARS = 40
MAX_ROWS_PER_FRAGMENT = 50  # keep fragments reasonably sized


def row_to_text(row) -> str:
    """Render one openpyxl row as a tab-separated string."""
    return " | ".join(
        str(cell.value).strip() if cell.value is not None else ""
        for cell in row
    ).rstrip(" | ")


def parse(path: str) -> list[dict]:
    source = Path(path)
    wb = load_workbook(source, read_only=True, data_only=True)
    fragments: list[dict] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        buffer: list[str] = []
        start_row = None
        current_row = 0

        for row in ws.iter_rows():
            current_row += 1
            row_text = row_to_text(row)
            if not row_text.strip():
                # Flush buffer on blank row gap
                if buffer:
                    _flush(fragments, source.name, sheet_name,
                           start_row, current_row - 1, buffer)
                    buffer = []
                    start_row = None
                continue

            if start_row is None:
                start_row = current_row
            buffer.append(row_text)

            # Flush when buffer reaches MAX size
            if len(buffer) >= MAX_ROWS_PER_FRAGMENT:
                _flush(fragments, source.name, sheet_name,
                       start_row, current_row, buffer)
                buffer = []
                start_row = None

        # Flush remaining rows in sheet
        if buffer:
            _flush(fragments, source.name, sheet_name,
                   start_row, current_row, buffer)

    wb.close()
    return fragments


def _flush(fragments, source_name, sheet_name, start_row, end_row, rows):
    text = "\n".join(rows).strip()
    if len(text) < MIN_CHARS:
        return
    fragments.append({
        "fragment_id": str(uuid.uuid4()),
        "source_file": source_name,
        "file_type": "xlsx",
        "location": f"Sheet '{sheet_name}' rows {start_row}–{end_row}",
        "text": text,
        "char_count": len(text),
    })


def main():
    if len(sys.argv) < 2:
        sys.exit("Usage: python parse_excel.py <file.xlsx>")

    for fragment in parse(sys.argv[1]):
        print(json.dumps(fragment))


if __name__ == "__main__":
    main()
